"""XLSX file reader for structured data."""

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterator, Any

import openpyxl

from .base import Reader
from .registry import reader_registry


@reader_registry.register_decorator("xlsx_file")
class XLSXReader(Reader):
    """
    Reader for Excel (XLSX) files.
    
    Reads XLSX files row-by-row, treating first row as headers.
    Each row becomes a dictionary with column names as keys.
    """
    
    def __init__(
        self,
        path: str | Path,
        config: Dict[str, Any] | None = None
    ):
        """
        Initialize XLSX reader.
        
        Args:
            path: Path to XLSX file
            config: Optional configuration dictionary
        """
        self.source_path = Path(path)
        
        # Extract XLSX-specific config
        xlsx_config = config.get("readers", {}).get("xlsx_file", {}) if config else {}
        self.sheet_name = xlsx_config.get("sheet_name", 0)  # 0 = first sheet
        self.skip_empty_rows = xlsx_config.get("skip_empty_rows", True)
        self.date_format = xlsx_config.get("date_format", "%Y-%m-%d")
    
    def _format_cell_value(self, value: Any) -> str:
        """
        Format cell value to string.
        
        Args:
            value: Cell value from openpyxl
            
        Returns:
            Formatted string value
        """
        if value is None:
            return ""
        
        # Handle datetime objects
        if isinstance(value, datetime):
            return value.strftime(self.date_format)
        
        # Convert everything else to string
        return str(value)
    
    def read(self) -> Iterator[Dict[str, Any]]:
        """
        Read XLSX file row-by-row.
        
        Yields:
            Dictionary for each row with metadata
        """
        if not self.source_path.exists():
            raise FileNotFoundError(f"XLSX file not found: {self.source_path}")
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(self.source_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load XLSX file: {e}")
        
        # Get the specified sheet
        if isinstance(self.sheet_name, int):
            # Use index (0-based)
            sheet_names = workbook.sheetnames
            if self.sheet_name >= len(sheet_names):
                raise ValueError(f"Sheet index {self.sheet_name} out of range. Available: {sheet_names}")
            sheet = workbook[sheet_names[self.sheet_name]]
            sheet_title = sheet_names[self.sheet_name]
        else:
            # Use sheet name
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {workbook.sheetnames}")
            sheet = workbook[self.sheet_name]
            sheet_title = self.sheet_name
        
        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return
        
        # First row is headers
        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
        
        # Process data rows
        for row_num, row_values in enumerate(rows[1:], start=2):  # Row 2 is first data row
            # Skip empty rows if configured
            if self.skip_empty_rows and all(v is None or str(v).strip() == "" for v in row_values):
                continue
            
            # Create dictionary with headers as keys
            row_data = {}
            for header, value in zip(headers, row_values):
                row_data[header] = self._format_cell_value(value)
            
            yield {
                "content": None,  # Will be populated by preprocessor
                "metadata": {
                    "source": str(self.source_path),
                    "sheet": sheet_title,
                    "row_number": row_num,
                    "original_data": row_data
                }
            }
