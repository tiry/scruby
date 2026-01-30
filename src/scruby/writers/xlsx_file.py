"""XLSX file writer for structured data."""

from pathlib import Path
from typing import Dict, Any

import openpyxl
from openpyxl import Workbook

from .base import Writer
from .registry import writer_registry


@writer_registry.register_decorator("xlsx_file")
class XLSXWriter(Writer):
    """
    Writer for Excel (XLSX) files.
    
    Writes dictionary data to Excel format, preserving column order.
    """
    
    def __init__(
        self,
        path: str | Path,
        config: Dict[str, Any] | None = None
    ):
        """
        Initialize XLSX writer.
        
        Args:
            path: Path to output XLSX file
            config: Optional configuration dictionary
        """
        self.destination_path = Path(path)
        
        # Extract XLSX-specific config
        xlsx_config = config.get("writers", {}).get("xlsx_file", {}) if config else {}
        self.sheet_name = xlsx_config.get("sheet_name", "Redacted Data")
        self.write_header = xlsx_config.get("write_header", True)
        
        # Initialize workbook
        self._workbook: Workbook | None = None
        self._worksheet = None
        self._fieldnames = None
        self._current_row = 1
    
    def write(self, document: Dict[str, Any]) -> None:
        """
        Write document as Excel row.
        
        Args:
            document: Document with redacted_data in metadata
        """
        # Get redacted data
        redacted_data = document.get("metadata", {}).get("redacted_data", {})
        
        if not redacted_data:
            return
        
        # Create workbook on first write
        if self._workbook is None:
            self._workbook = Workbook()
            self._worksheet = self._workbook.active
            self._worksheet.title = self.sheet_name
            
            # Get fieldnames from first document
            self._fieldnames = list(redacted_data.keys())
            
            # Write header if configured
            if self.write_header:
                for col_idx, fieldname in enumerate(self._fieldnames, start=1):
                    self._worksheet.cell(row=1, column=col_idx, value=fieldname)
                self._current_row = 2
        
        # Write data row
        for col_idx, fieldname in enumerate(self._fieldnames, start=1):
            value = redacted_data.get(fieldname, "")
            self._worksheet.cell(row=self._current_row, column=col_idx, value=value)
        
        self._current_row += 1
    
    def close(self) -> None:
        """Save and close the XLSX file."""
        if self._workbook:
            self.destination_path.parent.mkdir(parents=True, exist_ok=True)
            self._workbook.save(self.destination_path)
            self._workbook = None
            self._worksheet = None
    
    def __del__(self):
        """Ensure file is saved on deletion."""
        self.close()
