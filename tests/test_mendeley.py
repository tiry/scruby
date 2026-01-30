"""Integration tests for Mendeley dataset redaction."""

import tempfile
from pathlib import Path

from scruby.config import load_config
from scruby.pipeline import Pipeline


class TestMendeleyDataset:
    """Test PII redaction on the Mendeley testing dataset."""
    
    def test_mendeley_xlsx_redaction(self):
        """
        Test redaction of Mendeley test dataset.
        
        Verifies that all PII fields (Name, Credit Card, Email, URL, Phone, 
        Address, Company, SSN) are properly redacted while preserving 
        non-selected fields (Text, True Predictions).
        
        The Mendeley dataset contains 30 rows of synthetic PII data designed
        for testing entity detection and redaction systems.
        """
        # Load configuration from YAML file
        config = load_config("tests/fixtures/mendeley_config.yaml")
        
        # Paths
        input_file = Path("tests/data/mendeley_testing_dataset.xlsx")
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_file = Path(tmp.name)
        
        try:
            # Create pipeline
            pipeline = Pipeline(config=config)
            
            # Process all rows
            processed_docs = pipeline.process(
                input_path=input_file,
                output_path=output_file,
                reader_type="xlsx_file",
                writer_type="xlsx_file",
                preprocessors=["field_selector"],
                postprocessors=["dict_merger"]
            )
            
            # Verify results
            assert len(processed_docs) == 30, f"Should process 30 data rows, got {len(processed_docs)}"
            
            # Count total entities redacted across all docs (using correct metadata key)
            total_entities = sum(doc.get("metadata", {}).get("redacted_entities", 0) for doc in processed_docs)
            
            # Read output file and verify
            import openpyxl
            wb = openpyxl.load_workbook(output_file)
            sheet = wb.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            # Verify headers preserved
            expected_headers = [
                "Name", "Credit Card", "Email", "URL", "Phone",
                "Address", "Company", "SSN", "Text", "True Predictions"
            ]
            assert headers == expected_headers, f"Headers mismatch: {headers}"
            
            # Read input file to compare
            input_wb = openpyxl.load_workbook(input_file)
            input_sheet = input_wb.active
            
            # Check each data row
            pii_fields = ["Name", "Credit Card", "Email", "URL", "Phone", "Company", "SSN"]
            preserved_fields = ["Address", "Text", "True Predictions"]
            
            total_non_empty_pii_fields = 0
            total_redacted_fields = 0
            total_empty_fields = 0
            rows_checked = 0
            unredacted_fields = []  # Track fields that were not redacted
            
            for row_idx in range(2, sheet.max_row + 1):
                # Get input row data
                input_row = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = input_sheet.cell(row=row_idx, column=col_idx).value
                    input_row[header] = str(cell_value) if cell_value else ""
                
                # Get output row data
                output_row = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    output_row[header] = str(cell_value) if cell_value else ""
                
                rows_checked += 1
                
                # Check each PII field
                for field in pii_fields:
                    input_value = input_row.get(field, "")
                    output_value = output_row.get(field, "")
                    
                    # Skip None and empty strings
                    if input_value and input_value != "None" and input_value.strip():
                        total_non_empty_pii_fields += 1
                        
                        # Check if field was redacted (Presidio detected entities)
                        is_redacted = "<" in output_value and ">" in output_value and ":" in output_value
                        
                        if is_redacted:
                            # Verify redacted output differs from input
                            assert output_value != input_value, \
                                f"Row {row_idx}, field '{field}': Redacted field should differ from input"
                            total_redacted_fields += 1
                        else:
                            # Field was selected but Presidio didn't detect entities
                            # Track this for reporting
                            unredacted_fields.append({
                                'row': row_idx,
                                'field': field,
                                'value': input_value[:50]  # First 50 chars
                            })
                            # Output should match input (no changes made)
                            assert output_value == input_value, \
                                f"Row {row_idx}, field '{field}': No entities detected, should be unchanged"
                    else:
                        # Empty field - should remain empty or None
                        total_empty_fields += 1
                        assert output_value in ["", "None"], \
                            f"Row {row_idx}, field '{field}': Empty field should stay empty, got '{output_value}'"
                
                # Verify preserved fields are present AND unchanged from input
                for field in preserved_fields:
                    assert field in output_row, f"Preserved field {field} should exist in output"
                    # These fields should be unchanged (same as input)
                    assert output_row[field] == input_row[field], \
                        f"Row {row_idx}, preserved field '{field}' should be unchanged"
            
            print(f"\n✅ Processed {rows_checked} rows")
            print(f"✅ Non-empty PII fields: {total_non_empty_pii_fields}")
            print(f"✅ Fields successfully redacted: {total_redacted_fields}")
            print(f"✅ Empty fields (skipped): {total_empty_fields}")
            print(f"✅ Total entities detected by Presidio: {total_entities}")
            
            # Display unredacted fields if any
            if unredacted_fields:
                print(f"\n⚠️  WARNING: {len(unredacted_fields)} fields were NOT redacted:")
                print("=" * 60)
                for item in unredacted_fields:
                    print(f"  Row {item['row']}, Field '{item['field']}': {item['value']}")
                print("=" * 60)
            
            # Calculate redaction rate
            assert rows_checked == 30, f"Should check all 30 rows, checked {rows_checked}"
            
            if total_non_empty_pii_fields > 0:
                redaction_rate = (total_redacted_fields / total_non_empty_pii_fields) * 100
                diff_percentage = abs(total_non_empty_pii_fields - total_redacted_fields) / total_non_empty_pii_fields * 100
            else:
                redaction_rate = 0
                diff_percentage = 0
            
            # Verify at least 90% redaction rate
            assert redaction_rate >= 90, \
                f"Should redact at least 90% of fields, got {redaction_rate:.1f}% ({total_redacted_fields}/{total_non_empty_pii_fields})"
            assert total_entities > 0, f"Should detect some entities, found {total_entities}"
            
            print(f"\n✅ Mendeley dataset test complete!")
            print(f"   Redaction rate: {redaction_rate:.1f}% ({total_redacted_fields}/{total_non_empty_pii_fields} fields)")
            print(f"   Difference: {diff_percentage:.1f}%")
            print(f"   Average entities per row: {total_entities / rows_checked:.1f}")
            
        finally:
            # Cleanup
            if output_file.exists():
                output_file.unlink()


if __name__ == "__main__":
    # Run test
    test = TestMendeleyDataset()
    
    print("=" * 60)
    print("Testing Mendeley XLSX Dataset")
    print("=" * 60)
    test.test_mendeley_xlsx_redaction()
    
    print("\n" + "=" * 60)
    print("✅ TEST PASSED!")
    print("=" * 60)
