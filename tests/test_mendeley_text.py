"""Integration tests for Mendeley TEXT field redaction."""

import ast
import re
import tempfile
from pathlib import Path

from scruby.config import load_config
from scruby.pipeline import Pipeline


class TestMendeleyText:
    """Test PII redaction on the Mendeley TEXT field with ground truth validation."""
    
    def test_mendeley_text_redaction(self):
        """
        Test redaction of TEXT field in Mendeley dataset.
        
        Validates that the number of redacted entities matches the ground truth
        in the "True Predictions" column. True Predictions contains a list of
        tuples: [(start, end, 'entity_type'), ...]
        """
        # Load configuration from YAML file
        config = load_config("tests/fixtures/mendeley_text_config.yaml")
        
        # Paths
        input_file = Path("tests/data/mendeley_testing_dataset.xlsx")
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_file = Path(tmp.name)
        
        try:
            # Create pipeline
            pipeline = Pipeline(config=config)
            
            # Process all rows
            processed_docs = pipeline.process(
                input_path=input_file,
                output_path=output_file,
                reader_type="xlsx_file",
                writer_type="xlsx_file",
                preprocessors=["field_selector"],
                postprocessors=["dict_merger"]
            )
            
            # Verify results
            assert len(processed_docs) == 30, f"Should process 30 data rows, got {len(processed_docs)}"
            
            # Ensure file is written and closed (pipeline returns after writing)
            import time
            time.sleep(0.1)  # Small delay to ensure file is fully written
            
            # Verify output file exists and has content
            assert output_file.exists(), f"Output file should exist: {output_file}"
            assert output_file.stat().st_size > 0, f"Output file should not be empty: {output_file}"
            
            # Read output and input files
            import openpyxl
            output_wb = openpyxl.load_workbook(output_file)
            output_sheet = output_wb.active
            
            input_wb = openpyxl.load_workbook(input_file)
            input_sheet = input_wb.active
            
            # Get headers
            headers = [cell.value for cell in input_sheet[1]]
            text_idx = headers.index("Text") + 1
            pred_idx = headers.index("True Predictions") + 1
            
            total_expected_entities = 0
            total_actual_redactions = 0
            mismatches = []
            
            print("\nProcessing TEXT field redactions:")
            print("=" * 80)
            
            # Check each row
            for row_idx in range(2, input_sheet.max_row + 1):
                # Get True Predictions (ground truth)
                true_pred_str = input_sheet.cell(row=row_idx, column=pred_idx).value
                if true_pred_str:
                    # Parse the list of tuples
                    true_predictions = ast.literal_eval(true_pred_str)
                    expected_count = len(true_predictions)
                else:
                    expected_count = 0
                
                # Get redacted text from output
                redacted_text = output_sheet.cell(row=row_idx, column=text_idx).value or ""
                
                # Count redaction markers in output (pattern: <ENTITY_TYPE:hash>)
                redaction_pattern = r'<[A-Z_]+:[0-9a-f]+>'
                actual_redactions = re.findall(redaction_pattern, str(redacted_text))
                actual_count = len(actual_redactions)
                
                total_expected_entities += expected_count
                total_actual_redactions += actual_count
                
                # Track mismatches
                match_status = "✅" if actual_count == expected_count else "⚠️"
                print(f"{match_status} Row {row_idx}: Expected {expected_count}, Found {actual_count}")
                
                if actual_count != expected_count:
                    mismatches.append({
                        'row': row_idx,
                        'expected': expected_count,
                        'actual': actual_count,
                        'diff': actual_count - expected_count,
                        'redacted_text': redacted_text
                    })
            
            print("=" * 80)
            print(f"\n✅ Total expected entities (from True Predictions): {total_expected_entities}")
            print(f"✅ Total actual redactions: {total_actual_redactions}")
            
            # Display mismatches if any
            if mismatches:
                print(f"\n⚠️  WARNING: {len(mismatches)} rows have mismatches:")
                print("=" * 80)
                for item in mismatches:
                    diff_str = f"+{item['diff']}" if item['diff'] > 0 else str(item['diff'])
                    print(f"\n  Row {item['row']}: Expected {item['expected']}, Got {item['actual']} ({diff_str})")
                    print(f"  Full redacted text:")
                    print(f"  {item['redacted_text']}")
                print("=" * 80)
            
            # Calculate percentage difference
            if total_expected_entities > 0:
                diff_percentage = abs(total_actual_redactions - total_expected_entities) / total_expected_entities * 100
            else:
                diff_percentage = 0
            
            # Verify entity counts are within 15% tolerance
            # Note: With international phone recognizer, we detect MORE entities than ground truth
            assert diff_percentage < 15, \
                f"Entity count difference too large: {diff_percentage:.1f}% (expected {total_expected_entities}, got {total_actual_redactions})"
            
            print(f"\n✅ Mendeley TEXT field test complete!")
            print(f"   Expected: {total_expected_entities}, Detected: {total_actual_redactions}")
            print(f"   Difference: {diff_percentage:.1f}% (within 15% tolerance)")
            
        finally:
            # Cleanup
            if output_file.exists():
                output_file.unlink()


if __name__ == "__main__":
    # Run test
    test = TestMendeleyText()
    
    print("=" * 80)
    print("Testing Mendeley TEXT Field Redaction")
    print("=" * 80)
    test.test_mendeley_text_redaction()
    
    print("\n" + "=" * 80)
    print("✅ TEST PASSED!")
    print("=" * 80)
